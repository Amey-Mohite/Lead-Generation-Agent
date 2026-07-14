from pathlib import Path

import openpyxl

from app.exporters.excel import ExcelExporter
from app.schemas.lead import Lead, OutreachDraft, Qualification
from app.schemas.research import Contact, ResearchBrief

_HEADER = [
    "Company Name", "Domain", "Industry", "Status", "Score",
    "Qualification Reasoning", "Summary", "Key Facts", "Contacts",
    "Sources", "Outreach Subject", "Outreach Body", "Generated At",
]


def _qualified_lead() -> Lead:
    return Lead(
        research=ResearchBrief(
            company_name="Acme Credit Union",
            domain="acme-cu.com",
            industry="Financial Services",
            summary="A credit union.",
            key_facts=["Founded 1990", "10,000 members"],
            contacts=[Contact(name="Jane Doe", role="CTO", email="jane@acme-cu.com")],
            sources=["https://acme-cu.com"],
        ),
        qualification=Qualification(score=85, reasoning="Strong fit."),
        outreach=OutreachDraft(subject="Quick question", body="Hi Jane, ..."),
        status="qualified",
    )


def _disqualified_lead() -> Lead:
    return Lead(
        research=ResearchBrief(company_name="Beta Corp", summary="Not a fit."),
        qualification=Qualification(score=20, reasoning="Wrong industry."),
        outreach=None,
        status="disqualified",
    )


def test_export_writes_expected_header_and_rows(tmp_path):
    exporter = ExcelExporter(export_dir=str(tmp_path))
    leads = [_qualified_lead(), _disqualified_lead()]

    path = exporter.export(leads)

    assert path.endswith(".xlsx")
    wb = openpyxl.load_workbook(path)
    ws = wb.active

    header = [cell.value for cell in ws[1]]
    assert header == _HEADER

    row1 = [cell.value for cell in ws[2]]
    assert row1[0] == "Acme Credit Union"
    assert row1[3] == "qualified"
    assert row1[4] == 85
    assert "Founded 1990" in row1[7]
    assert "Jane Doe (CTO) <jane@acme-cu.com>" in row1[8]
    assert row1[10] == "Quick question"
    assert row1[12] is not None  # Generated At populated

    row2 = [cell.value for cell in ws[3]]
    assert row2[0] == "Beta Corp"
    assert row2[3] == "disqualified"
    assert row2[10] is None  # no outreach subject
    assert row2[11] is None  # no outreach body


def test_export_creates_export_dir_if_missing(tmp_path):
    export_dir = tmp_path / "nested" / "dir"
    exporter = ExcelExporter(export_dir=str(export_dir))

    path = exporter.export([_qualified_lead()])

    assert Path(path).exists()


def test_export_writes_a_fresh_timestamped_file_each_call(tmp_path):
    exporter = ExcelExporter(export_dir=str(tmp_path))

    path1 = exporter.export([_qualified_lead()])
    path2 = exporter.export([_qualified_lead()])

    assert path1 != path2
    assert Path(path1).exists()
    assert Path(path2).exists()
